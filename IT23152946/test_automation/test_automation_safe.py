from playwright.sync_api import sync_playwright
from pathlib import Path
import argparse, time, re, sys, os
import openpyxl
from openpyxl.cell.cell import MergedCell

DEFAULT_URL = "https://www.pixelssuite.com/chat-translator"

def norm(x):
    return re.sub(r"[^a-z0-9]+", "", str(x or "").strip().lower())

def configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

def top_left_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return cell

def is_top_left(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def set_value(ws, row, col, value):
    top_left_cell(ws, row, col).value = value

def find_header(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        vals = [norm(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "tcid" in vals and "input" in vals and "expectedoutput" in vals:
            return r
    return 1

def find_col(ws, header_row, names):
    wanted = {norm(n) for n in names}
    for c in range(1, ws.max_column + 1):
        if norm(ws.cell(header_row, c).value) in wanted:
            return c
    # partial fallback
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(header_row, c).value)
        for w in wanted:
            if h and (h in w or w in h):
                return c
    return None

def ensure_col(ws, header_row, name):
    c = find_col(ws, header_row, [name])
    if c:
        return c
    c = ws.max_column + 1
    ws.cell(header_row, c).value = name
    return c

def read_output(locator):
    for method in ("input_value", "inner_text", "text_content"):
        try:
            v = getattr(locator, method)()
            if v and str(v).strip():
                return str(v).strip()
        except Exception:
            pass
    try:
        v = locator.evaluate("(el) => el && ('value' in el ? el.value : el.textContent)")
        if v and str(v).strip():
            return str(v).strip()
    except Exception:
        pass
    return ""

def find_ui(page, timeout_ms):
    page.wait_for_selector("textarea", timeout=timeout_ms)
    input_box = page.locator('textarea[placeholder*="English"]').first
    output_box = page.locator('textarea[placeholder*="Sinhala"]').first
    if input_box.count() == 0 or output_box.count() == 0:
        areas = page.locator("textarea")
        input_box = areas.nth(0)
        output_box = areas.nth(1)
    button = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.I)).first
    return input_box, output_box, button

def clear_and_type(page, box, text, type_delay):
    box.click(timeout=5000)
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    box.fill("")
    if type_delay > 0:
        box.type(text, delay=type_delay)
    else:
        box.fill(text)

def main():
    configure_stdout()
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default="Assignment 1 - Test cases.xlsx")
    parser.add_argument("--url", default=DEFAULT_URL)
    parser.add_argument("--wait-ms", type=int, default=90000)
    parser.add_argument("--timeout-ms", type=int, default=180000)
    parser.add_argument("--type-delay-ms", type=int, default=80)
    parser.add_argument("--between-ms", type=int, default=10000)
    parser.add_argument("--reload-each", action="store_true", default=True)
    parser.add_argument("--headless", action="store_true", default=False)
    args = parser.parse_args()

    excel_path = Path(args.excel).resolve()
    if not excel_path.exists():
        print(f"Excel file not found: {excel_path}")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[" Test cases"] if " Test cases" in wb.sheetnames else wb.active
    header = find_header(ws)

    input_col = find_col(ws, header, ["Input", "Singlish Input", "Test Input"])
    expected_col = find_col(ws, header, ["Expected output", "Expected Output", "Expected Sinhala"])
    actual_col = ensure_col(ws, header, "Actual output")
    status_col = ensure_col(ws, header, "Status")

    if not input_col:
        print("Input column not found.")
        return

    rows = []
    for r in range(header + 1, ws.max_row + 1):
        if not is_top_left(ws, r, input_col):
            continue
        txt = top_left_cell(ws, r, input_col).value
        if txt and str(txt).strip():
            # skip only successful collected rows; rerun blank/UI Error rows
            actual = str(top_left_cell(ws, r, actual_col).value or "").strip()
            status = str(top_left_cell(ws, r, status_col).value or "").strip().upper()
            if actual and status not in ("UI ERROR", "ERROR"):
                print(f"Skipping Row {r}: already has actual output")
                continue
            rows.append((r, str(txt).strip()))

    print(f"Rows to test: {len(rows)}")
    print("Important: Do not click the browser while running.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=args.headless, slow_mo=150)
        page = browser.new_page()
        page.set_default_timeout(args.timeout_ms)

        for index, (row, text) in enumerate(rows, 1):
            print(f"\n[{index}/{len(rows)}] Row {row}: {text}")

            try:
                page.goto(args.url, wait_until="domcontentloaded", timeout=args.timeout_ms)
                try:
                    page.wait_for_load_state("networkidle", timeout=30000)
                except Exception:
                    pass

                input_box, output_box, button = find_ui(page, args.timeout_ms)
                clear_and_type(page, input_box, text, args.type_delay_ms)

                before = read_output(output_box)
                button.click(timeout=30000)

                actual = ""
                start = time.time()
                while (time.time() - start) * 1000 < args.wait_ms:
                    current = read_output(output_box)
                    if current and current != before and "failed to fetch" not in current.lower():
                        actual = current
                        break
                    if current and "failed to fetch" in current.lower():
                        # wait more; sometimes the next call works after reload, but record UI Error for this row
                        actual = current
                        break
                    page.wait_for_timeout(2000)

                if not actual:
                    set_value(ws, row, actual_col, "")
                    set_value(ws, row, status_col, "UI Error")
                    print("  -> UI Error (no actual output)")
                elif "failed to fetch" in actual.lower():
                    set_value(ws, row, actual_col, actual)
                    set_value(ws, row, status_col, "UI Error")
                    print("  -> UI Error (Failed to fetch)")
                else:
                    expected = str(top_left_cell(ws, row, expected_col).value or "").strip() if expected_col else ""
                    set_value(ws, row, actual_col, actual)
                    set_value(ws, row, status_col, "PASS" if expected and actual == expected else "FAIL")
                    print(f"  -> saved: {str(actual)[:80]}")

                wb.save(excel_path)
                page.wait_for_timeout(args.between_ms)

            except Exception as e:
                print(f"  -> UI Error: {e}")
                set_value(ws, row, status_col, "UI Error")
                wb.save(excel_path)
                try:
                    page.wait_for_timeout(args.between_ms)
                except Exception:
                    pass

        browser.close()

    wb.save(excel_path)
    print(f"\nDone. Results saved to: {excel_path}")

if __name__ == "__main__":
    main()
